import streamlit as st
import pandas as pd
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import io
import time
import numpy as np
from datetime import datetime
import math
import logging
from typing import Tuple, List, Optional
import warnings

# Gestion des imports optionnels
try:
    from sklearn.cluster import DBSCAN
    from sklearn.preprocessing import StandardScaler
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Supprimer les warnings pandas
warnings.filterwarnings('ignore', category=FutureWarning)

# Configuration de la page
st.set_page_config(
    page_title="Organisateur de Tournées Automatique",
    page_icon="🚛",
    layout="wide"
)

# Titre principal
st.title("🚛 Organisateur de Tournées Automatique")
st.markdown("*Optimisation intelligente pour itinéraires de livraison*")
st.markdown("---")

@st.cache_data
def geocode_address(address: str, max_retries: int = 3) -> Tuple[Optional[float], Optional[float], bool]:
    """
    Géocode une adresse avec retry intelligent et gestion d'erreurs améliorée
    
    Args:
        address (str): L'adresse à géocoder
        max_retries (int): Nombre maximum de tentatives
        
    Returns:
        Tuple[Optional[float], Optional[float], bool]: (latitude, longitude, success)
    """
    geolocator = Nominatim(user_agent="delivery_route_optimizer_v2", timeout=20)
    
    for attempt in range(max_retries):
        try:
            # Attendre entre les tentatives avec backoff exponentiel
            if attempt > 0:
                wait_time = min(2 ** attempt, 8)  # Maximum 8 secondes
                time.sleep(wait_time)
            
            location = geolocator.geocode(address, timeout=15)
            if location:
                logger.info(f"Geocoding successful for: {address[:50]}...")
                return (location.latitude, location.longitude, True)
            else:
                logger.warning(f"No location found for: {address[:50]}...")
                return (None, None, False)
                
        except GeocoderTimedOut:
            logger.warning(f"Geocoding timeout for: {address[:50]}... (attempt {attempt + 1}/{max_retries})")
            if attempt == max_retries - 1:
                return (None, None, False)
            
        except GeocoderServiceError as e:
            logger.error(f"Geocoding service error for: {address[:50]}... - {str(e)}")
            if attempt == max_retries - 1:
                return (None, None, False)
            
        except Exception as e:
            logger.error(f"Unexpected geocoding error for: {address[:50]}... - {str(e)}")
            if attempt == max_retries - 1:
                return (None, None, False)
    
    return (None, None, False)

def find_center_city_point(df_points: pd.DataFrame) -> int:
    """
    Trouve le point central de la ville (zone avec la plus forte densité de livraisons)
    
    Args:
        df_points (pd.DataFrame): DataFrame avec les coordonnées des points
        
    Returns:
        int: Index du point central
    """
    if len(df_points) < 3:
        return df_points.iloc[0].name
    
    # Calcul de la densité pour chaque point
    density_scores = []
    
    for idx, row in df_points.iterrows():
        current_coords = (row['lat'], row['lon'])
        
        # Compter les points dans un rayon de 2km et calculer la densité pondérée
        density_score = 0
        total_weight = 0
        
        for idx2, row2 in df_points.iterrows():
            if idx != idx2:
                point_coords = (row2['lat'], row2['lon'])
                distance = geodesic(current_coords, point_coords).kilometers
                
                # Pondération inversement proportionnelle à la distance
                if distance <= 3.0:  # Rayon étendu à 3km
                    weight = 1 / (distance + 0.1)  # Éviter division par 0
                    density_score += weight
                    total_weight += 1
        
        # Score final normalisé
        final_score = density_score / max(total_weight, 1)
        density_scores.append((idx, final_score, current_coords))
    
    # Trier par densité décroissante
    density_scores.sort(key=lambda x: x[1], reverse=True)
    
    # Retourner l'index du point avec la plus forte densité
    return density_scores[0][0]

def filter_addresses_by_proximity(df_points: pd.DataFrame, max_radius_km: float = 15) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filtre les adresses pour garder seulement celles dans un secteur cohérent
    
    Args:
        df_points (pd.DataFrame): DataFrame avec les coordonnées des points
        max_radius_km (float): Rayon maximum en kilomètres
        
    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: (adresses dans le secteur, adresses hors secteur)
    """
    if len(df_points) < 3:
        return df_points, pd.DataFrame()
    
    try:
        # Étape 1: Trouver le centre de distribution (zone dense)
        center_idx = find_center_city_point(df_points)
        center_coords = (df_points.at[center_idx, 'lat'], df_points.at[center_idx, 'lon'])
        
        # Étape 2: Calculer les distances depuis le centre
        distances_from_center = []
        for idx, row in df_points.iterrows():
            point_coords = (row['lat'], row['lon'])
            distance = geodesic(center_coords, point_coords).kilometers
            distances_from_center.append((idx, distance))
        
        # Étape 3: Analyse de la distribution des distances
        distances_only = [d[1] for d in distances_from_center]
        
        # Calcul automatique du rayon optimal avec méthode améliorée
        if len(distances_only) > 10:
            # Méthode des quantiles pour détecter les valeurs aberrantes
            q75 = np.percentile(distances_only, 75)
            q25 = np.percentile(distances_only, 25)
            iqr = q75 - q25
            
            # Seuil adaptatif avec prise en compte de la médiane
            median_dist = np.median(distances_only)
            max_distance = min(max_radius_km, max(q75 + 1.5 * iqr, median_dist * 2))
        else:
            # Pour de petits échantillons, utiliser la médiane + écart-type
            median_dist = np.median(distances_only)
            std_dist = np.std(distances_only)
            max_distance = min(max_radius_km, median_dist + 1.5 * std_dist)
        
        # Assurer un minimum de 3km et maximum de max_radius_km
        max_distance = max(3.0, min(max_distance, max_radius_km))
        
        # Étape 4: Filtrage des adresses
        addresses_in_sector = []
        addresses_out_of_sector = []
        
        for idx, distance in distances_from_center:
            if distance <= max_distance:
                addresses_in_sector.append(idx)
            else:
                addresses_out_of_sector.append(idx)
        
        df_in_sector = df_points.loc[addresses_in_sector].copy()
        df_out_of_sector = df_points.loc[addresses_out_of_sector].copy()
        
        logger.info(f"Filtered addresses: {len(df_in_sector)} in sector, {len(df_out_of_sector)} out of sector")
        
        return df_in_sector, df_out_of_sector
        
    except Exception as e:
        logger.error(f"Error in filter_addresses_by_proximity: {str(e)}")
        return df_points, pd.DataFrame()

def optimize_delivery_route_from_center(df_points: pd.DataFrame) -> List[int]:
    """
    Optimisation de tournée qui commence par le centre-ville avec algorithme amélioré
    
    Args:
        df_points (pd.DataFrame): DataFrame avec les coordonnées des points
        
    Returns:
        List[int]: Liste des indices dans l'ordre optimal
    """
    if len(df_points) <= 1:
        return df_points.index.tolist()
    
    try:
        # Étape 1: Identifier le point de départ (centre-ville)
        start_idx = find_center_city_point(df_points)
        
        # Étape 2: Créer des clusters/zones pour organiser la tournée
        route_order = create_smart_route_from_center(df_points, start_idx)
        
        # Étape 3: Optimisation locale avec 2-opt amélioré
        route_order = improve_route_2opt(df_points, route_order)
        
        # Étape 4: Optimisation finale avec heuristique du plus proche voisin
        route_order = refine_route_nearest_neighbor(df_points, route_order)
        
        logger.info(f"Route optimization completed: {len(route_order)} points")
        return route_order
        
    except Exception as e:
        logger.error(f"Error in optimize_delivery_route_from_center: {str(e)}")
        return df_points.index.tolist()

def create_smart_route_from_center(df_points, start_idx):
    """
    Crée un itinéraire intelligent qui part du centre et suit une logique géographique
    """
    if len(df_points) <= 2:
        return df_points.index.tolist()
    
    # Point de départ
    start_coords = (df_points.at[start_idx, 'lat'], df_points.at[start_idx, 'lon'])
    
    # Organiser les points par direction/angle depuis le centre
    points_with_angles = []
    
    for idx, row in df_points.iterrows():
        if idx == start_idx:
            continue
            
        point_coords = (row['lat'], row['lon'])
        
        # Calculer l'angle par rapport au centre
        lat_diff = point_coords[0] - start_coords[0]
        lon_diff = point_coords[1] - start_coords[1]
        angle = math.atan2(lat_diff, lon_diff)
        
        # Calculer la distance
        distance = geodesic(start_coords, point_coords).kilometers
        
        points_with_angles.append((idx, angle, distance))
    
    # Trier par angle pour créer un parcours circulaire
    points_with_angles.sort(key=lambda x: x[1])
    
    # Créer l'itinéraire : départ du centre, puis parcours circulaire
    route = [start_idx]
    
    # Ajouter les points dans l'ordre angulaire, en privilégiant les plus proches d'abord
    remaining_points = points_with_angles.copy()
    
    while remaining_points:
        # Trouver le point le plus proche parmi ceux restants
        current_coords = (df_points.at[route[-1], 'lat'], df_points.at[route[-1], 'lon'])
        
        best_idx = None
        best_distance = float('inf')
        best_position = None
        
        for i, (idx, angle, dist_from_center) in enumerate(remaining_points):
            point_coords = (df_points.at[idx, 'lat'], df_points.at[idx, 'lon'])
            distance_to_current = geodesic(current_coords, point_coords).kilometers
            
            # Pondérer par la distance au point actuel et la distance au centre
            # Privilégier les points proches du point actuel
            score = distance_to_current + (dist_from_center * 0.1)
            
            if score < best_distance:
                best_distance = score
                best_idx = idx
                best_position = i
        
        route.append(best_idx)
        remaining_points.pop(best_position)
    
    return route

def refine_route_nearest_neighbor(df_points: pd.DataFrame, route: List[int]) -> List[int]:
    """
    Affine la route avec une heuristique du plus proche voisin pour les segments problématiques
    
    Args:
        df_points (pd.DataFrame): DataFrame avec les coordonnées des points
        route (List[int]): Route actuelle
        
    Returns:
        List[int]: Route affinée
    """
    if len(route) <= 3:
        return route
    
    def calculate_segment_distance(idx1: int, idx2: int) -> float:
        coord1 = (df_points.at[idx1, 'lat'], df_points.at[idx1, 'lon'])
        coord2 = (df_points.at[idx2, 'lat'], df_points.at[idx2, 'lon'])
        return geodesic(coord1, coord2).kilometers
    
    # Identifier les segments les plus longs (potentiellement problématiques)
    segment_distances = []
    for i in range(len(route) - 1):
        distance = calculate_segment_distance(route[i], route[i + 1])
        segment_distances.append((i, distance))
    
    # Trier par distance décroissante
    segment_distances.sort(key=lambda x: x[1], reverse=True)
    
    # Optimiser les 20% des segments les plus longs
    num_segments_to_optimize = max(1, len(segment_distances) // 5)
    
    for i in range(num_segments_to_optimize):
        segment_idx = segment_distances[i][0]
        if segment_idx < len(route) - 1:
            # Essayer de trouver un meilleur point intermédiaire
            current_point = route[segment_idx]
            next_point = route[segment_idx + 1]
            
            # Chercher dans les points environnants
            best_improvement = 0
            best_point = None
            
            for j in range(max(0, segment_idx - 2), min(len(route), segment_idx + 4)):
                if j != segment_idx and j != segment_idx + 1:
                    test_point = route[j]
                    
                    # Calculer l'amélioration potentielle
                    original_distance = calculate_segment_distance(current_point, next_point)
                    new_distance = (calculate_segment_distance(current_point, test_point) + 
                                  calculate_segment_distance(test_point, next_point))
                    
                    improvement = original_distance - new_distance
                    if improvement > best_improvement:
                        best_improvement = improvement
                        best_point = test_point
            
            # Appliquer l'amélioration si elle est significative
            if best_improvement > 0.5:  # Amélioration d'au moins 500m
                # Réorganiser la route
                new_route = route.copy()
                point_idx = new_route.index(best_point)
                new_route.pop(point_idx)
                new_route.insert(segment_idx + 1, best_point)
                route = new_route
    
    return route
def improve_route_2opt(df_points: pd.DataFrame, route: List[int]) -> List[int]:
    """Amélioration de la route avec l'algorithme 2-opt amélioré"""
    if len(route) < 4:
        return route
    
    def calculate_route_distance(route_order: List[int]) -> float:
        total_dist = 0
        for i in range(len(route_order) - 1):
            coord1 = (df_points.at[route_order[i], 'lat'], df_points.at[route_order[i], 'lon'])
            coord2 = (df_points.at[route_order[i+1], 'lat'], df_points.at[route_order[i+1], 'lon'])
            total_dist += geodesic(coord1, coord2).kilometers
        return total_dist
    
    best_route = route.copy()
    best_distance = calculate_route_distance(best_route)
    
    improved = True
    iterations = 0
    max_iterations = min(100, len(route) * 2)  # Augmenter les itérations pour de meilleurs résultats
    
    while improved and iterations < max_iterations:
        improved = False
        iterations += 1
        
        # Ne pas modifier le premier point (centre-ville)
        for i in range(1, len(route) - 2):
            for j in range(i + 2, len(route)):
                if j - i == 1:
                    continue
                
                # Créer nouvelle route en inversant le segment
                new_route = route.copy()
                new_route[i:j] = route[i:j][::-1]
                
                new_distance = calculate_route_distance(new_route)
                
                if new_distance < best_distance:
                    best_route = new_route
                    best_distance = new_distance
                    route = new_route
                    improved = True
                    break
            
            if improved:
                break
    
    logger.info(f"2-opt optimization completed in {iterations} iterations")
    return best_route

def calculate_route_stats(df_route):
    """Calcul des statistiques de la tournée"""
    if len(df_route) < 2:
        return 0, 0, 0
    
    total_distance = 0
    distances = []
    
    for i in range(len(df_route) - 1):
        coord1 = (df_route.iloc[i]['lat'], df_route.iloc[i]['lon'])
        coord2 = (df_route.iloc[i + 1]['lat'], df_route.iloc[i + 1]['lon'])
        segment_distance = geodesic(coord1, coord2).kilometers
        total_distance += segment_distance
        distances.append(segment_distance)
    
    # Statistiques
    avg_distance = np.mean(distances) if distances else 0
    max_distance = max(distances) if distances else 0
    
    return total_distance, avg_distance, max_distance

def detect_columns_smart(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Détection intelligente des colonnes adresse, code postal et ville avec algorithme amélioré
    
    Args:
        df (pd.DataFrame): DataFrame à analyser
        
    Returns:
        Tuple[Optional[str], Optional[str], Optional[str]]: (colonne_adresse, colonne_postal, colonne_ville)
    """
    columns = df.columns.tolist()
    
    # Mots-clés pour la détection avec priorités
    address_keywords = {
        'high': ['adresse', 'address', 'client', 'livraison'],
        'medium': ['rue', 'street', 'voie', 'avenue', 'boulevard', 'chemin', 'route'],
        'low': ['addr', 'add', 'lieu', 'location']
    }
    
    postal_keywords = {
        'high': ['postal', 'cp', 'code'],
        'medium': ['zip', 'postcode', 'post_code'],
        'low': ['cd', 'pc']
    }
    
    city_keywords = {
        'high': ['ville', 'city', 'commune'],
        'medium': ['localite', 'locality', 'town'],
        'low': ['municipalite', 'place']
    }
    
    def score_column(col_name: str, keywords_dict: dict) -> float:
        """Calcule un score pour une colonne basé sur les mots-clés"""
        col_lower = col_name.lower()
        score = 0
        
        for priority, keywords in keywords_dict.items():
            weight = {'high': 3, 'medium': 2, 'low': 1}[priority]
            for keyword in keywords:
                if keyword in col_lower:
                    score += weight
                    # Bonus si le mot-clé correspond exactement
                    if keyword == col_lower:
                        score += 2
        
        return score
    
    # Analyse des données pour confirmer les types
    def analyze_column_content(col_name: str) -> dict:
        """Analyse le contenu d'une colonne pour déterminer son type"""
        try:
            sample = df[col_name].dropna().head(20).astype(str)
            analysis = {
                'avg_length': sample.str.len().mean(),
                'has_numbers': sample.str.contains(r'\d').sum(),
                'has_spaces': sample.str.contains(' ').sum(),
                'numeric_ratio': sample.str.isdigit().sum() / len(sample)
            }
            return analysis
        except:
            return {'avg_length': 0, 'has_numbers': 0, 'has_spaces': 0, 'numeric_ratio': 0}
    
    # Scoring des colonnes
    address_scores = []
    postal_scores = []
    city_scores = []
    
    for col in columns:
        # Score basé sur les mots-clés
        addr_score = score_column(col, address_keywords)
        postal_score = score_column(col, postal_keywords)
        city_score = score_column(col, city_keywords)
        
        # Analyse du contenu
        content_analysis = analyze_column_content(col)
        
        # Ajustement des scores basé sur le contenu
        if content_analysis['avg_length'] > 20:  # Probablement une adresse
            addr_score += 1
        if content_analysis['numeric_ratio'] > 0.8:  # Probablement un code postal
            postal_score += 2
        if 5 <= content_analysis['avg_length'] <= 30 and content_analysis['has_spaces'] / len(df[col].dropna().head(20)) > 0.3:
            city_score += 1
        
        address_scores.append((col, addr_score))
        postal_scores.append((col, postal_score))
        city_scores.append((col, city_score))
    
    # Sélection des meilleures colonnes
    address_scores.sort(key=lambda x: x[1], reverse=True)
    postal_scores.sort(key=lambda x: x[1], reverse=True)
    city_scores.sort(key=lambda x: x[1], reverse=True)
    
    address_col = address_scores[0][0] if address_scores[0][1] > 0 else None
    postal_col = postal_scores[0][0] if postal_scores[0][1] > 0 else None
    city_col = city_scores[0][0] if city_scores[0][1] > 0 else None
    
    # Éviter les doublons
    used_columns = set()
    final_cols = [None, None, None]
    
    for i, (col, score_list) in enumerate([(address_col, address_scores), (postal_col, postal_scores), (city_col, city_scores)]):
        if col and col not in used_columns:
            final_cols[i] = col
            used_columns.add(col)
        else:
            # Chercher la meilleure alternative
            for alt_col, alt_score in score_list:
                if alt_col not in used_columns and alt_score > 0:
                    final_cols[i] = alt_col
                    used_columns.add(alt_col)
                    break
    
    logger.info(f"Column detection: address={final_cols[0]}, postal={final_cols[1]}, city={final_cols[2]}")
    
    return final_cols[0], final_cols[1], final_cols[2]

def validate_input_data(df: pd.DataFrame, address_col: str, postal_col: str, city_col: str) -> Tuple[bool, str]:
    """
    Valide les données d'entrée
    
    Args:
        df (pd.DataFrame): DataFrame à valider
        address_col, postal_col, city_col (str): Noms des colonnes
        
    Returns:
        Tuple[bool, str]: (is_valid, error_message)
    """
    try:
        # Vérifier que les colonnes existent
        missing_cols = []
        for col in [address_col, postal_col, city_col]:
            if col not in df.columns:
                missing_cols.append(col)
        
        if missing_cols:
            return False, f"Colonnes manquantes: {', '.join(missing_cols)}"
        
        # Vérifier qu'il y a des données
        if len(df) == 0:
            return False, "Le fichier ne contient aucune donnée"
        
        # Vérifier que les colonnes ne sont pas entièrement vides
        for col in [address_col, postal_col, city_col]:
            if df[col].isna().all():
                return False, f"La colonne '{col}' est entièrement vide"
        
        # Vérifier qu'il y a au moins quelques lignes complètes
        complete_rows = df.dropna(subset=[address_col, postal_col, city_col])
        if len(complete_rows) < 2:
            return False, "Trop peu de lignes complètes (minimum 2 requis)"
        
        return True, "Données valides"
        
    except Exception as e:
        return False, f"Erreur de validation: {str(e)}"

def calculate_route_quality_metrics(df_route: pd.DataFrame) -> dict:
    """
    Calcule des métriques de qualité pour la route
    
    Args:
        df_route (pd.DataFrame): DataFrame de la route optimisée
        
    Returns:
        dict: Métriques de qualité
    """
    if len(df_route) < 2:
        return {}
    
    try:
        # Calcul des distances
        distances = []
        for i in range(len(df_route) - 1):
            coord1 = (df_route.iloc[i]['lat'], df_route.iloc[i]['lon'])
            coord2 = (df_route.iloc[i + 1]['lat'], df_route.iloc[i + 1]['lon'])
            dist = geodesic(coord1, coord2).kilometers
            distances.append(dist)
        
        # Métriques de base
        total_distance = sum(distances)
        avg_distance = np.mean(distances)
        std_distance = np.std(distances)
        max_distance = max(distances)
        min_distance = min(distances)
        
        # Métriques de qualité
        # Coefficient de variation (plus faible = plus régulier)
        cv = std_distance / avg_distance if avg_distance > 0 else 0
        
        # Efficacité (ratio entre distance directe et distance totale)
        if len(df_route) >= 2:
            direct_distance = geodesic(
                (df_route.iloc[0]['lat'], df_route.iloc[0]['lon']),
                (df_route.iloc[-1]['lat'], df_route.iloc[-1]['lon'])
            ).kilometers
            efficiency = direct_distance / total_distance if total_distance > 0 else 0
        else:
            efficiency = 1.0
        
        # Compacité (mesure de la dispersion des points)
        center_lat = df_route['lat'].mean()
        center_lon = df_route['lon'].mean()
        distances_from_center = [
            geodesic((center_lat, center_lon), (row['lat'], row['lon'])).kilometers
            for _, row in df_route.iterrows()
        ]
        compactness = np.std(distances_from_center)
        
        return {
            'total_distance': total_distance,
            'avg_distance': avg_distance,
            'std_distance': std_distance,
            'max_distance': max_distance,
            'min_distance': min_distance,
            'coefficient_variation': cv,
            'efficiency': efficiency,
            'compactness': compactness,
            'quality_score': (1 - cv) * efficiency * (1 / (1 + compactness))  # Score composite
        }
        
    except Exception as e:
        logger.error(f"Error calculating route quality metrics: {str(e)}")
        return {}

# Interface utilisateur
def main():
    """Main function containing the Streamlit interface"""
    st.header("📁 Chargement du fichier")
    uploaded_file = st.file_uploader(
        "Déposez votre fichier Excel avec les adresses clients",
        type=["xlsx"],
        help="Le fichier doit contenir les colonnes : adresse, code postal, ville"
    )

    if not uploaded_file:
        st.info("🔄 En attente du fichier Excel...")
        st.markdown("### 📋 Format attendu :")
        st.markdown("- **Colonne 1:** Adresse du client")
        st.markdown("- **Colonne 2:** Code postal")
        st.markdown("- **Colonne 3:** Ville")
        st.stop()

    # Lecture du fichier
    try:
        df = pd.read_excel(uploaded_file)
        st.success(f"✅ Fichier chargé avec succès : {len(df)} adresses")
    except Exception as e:
        st.error(f"❌ Erreur lors de la lecture du fichier : {e}")
        st.stop()

    # Aperçu des données
    st.markdown("---")
    st.subheader("📊 Aperçu des données")
    st.dataframe(df.head(10), use_container_width=True)

# Détection automatique intelligente
address_col, postal_col, city_col = detect_columns_smart(df)

# Si détection automatique échoue, permettre sélection manuelle
if not all([address_col, postal_col, city_col]):
    st.warning("⚠️ Détection automatique des colonnes échouée. Sélection manuelle requise.")
    
    col1, col2, col3 = st.columns(3)
    columns = df.columns.tolist()
    
    with col1:
        address_col = st.selectbox("📍 Colonne Adresse", columns, key="manual_address")
    with col2:
        postal_col = st.selectbox("📮 Colonne Code Postal", columns, key="manual_postal")
    with col3:
        city_col = st.selectbox("🏙️ Colonne Ville", columns, key="manual_city")
    
    if not all([address_col, postal_col, city_col]):
        st.error("❌ Veuillez sélectionner toutes les colonnes requises")
        st.stop()
    
    st.info(f"📍 Colonnes sélectionnées : {address_col}, {postal_col}, {city_col}")
    
    # Validation des données
    is_valid, validation_message = validate_input_data(df, address_col, postal_col, city_col)
    if not is_valid:
        st.error(f"❌ Erreur de validation : {validation_message}")
        st.stop()
    else:
        st.success(f"✅ Validation réussie : {validation_message}")
else:
    st.success(f"✅ Colonnes détectées automatiquement : {address_col}, {postal_col}, {city_col}")
    
    # Validation des données détectées automatiquement
    is_valid, validation_message = validate_input_data(df, address_col, postal_col, city_col)
    if not is_valid:
        st.error(f"❌ Erreur de validation : {validation_message}")
        st.stop()

# Paramètres de filtrage
st.markdown("---")
st.subheader("⚙️ Paramètres de la tournée")

col1, col2 = st.columns(2)
with col1:
    max_radius = st.slider(
        "🎯 Rayon maximum de livraison (km)",
        min_value=5,
        max_value=25,
        value=15,
        help="Distance maximale depuis le centre-ville"
    )
with col2:
    max_addresses = st.slider(
        "📦 Nombre maximum d'adresses",
        min_value=10,
        max_value=200,
        value=100,
        help="Limite pour éviter les quotas API"
    )

# Validation et nettoyage des données
st.markdown("---")
st.subheader("🧹 Nettoyage des données")

# Nettoyage des données
df_clean = df.dropna(subset=[address_col, postal_col, city_col]).copy()

# Nettoyage des codes postaux
df_clean[postal_col] = df_clean[postal_col].astype(str).str.extract('(\d+)')[0]
df_clean = df_clean.dropna(subset=[postal_col])

# Suppression des lignes avec des valeurs vides
df_clean = df_clean[df_clean[address_col].astype(str).str.strip() != '']
df_clean = df_clean[df_clean[city_col].astype(str).str.strip() != '']
df_clean = df_clean[df_clean[postal_col].astype(str).str.strip() != '']

# Limitation
if len(df_clean) > max_addresses:
    st.warning(f"⚠️ Limitation à {max_addresses} adresses pour éviter les quotas API")
    df_clean = df_clean.head(max_addresses)

st.info(f"📊 Données nettoyées : {len(df_clean)} adresses valides")

if len(df_clean) == 0:
    st.error("❌ Aucune donnée valide après nettoyage")
    st.stop()

# Bouton de traitement
if st.button("🚀 Organiser la tournée intelligente", type="primary", use_container_width=True):
    
    # Construction des adresses complètes
    df_clean['adresse_complete'] = (
        df_clean[address_col].astype(str) + ", " + 
        df_clean[postal_col].astype(str) + " " + 
        df_clean[city_col].astype(str) + ", France"
    )
    
    # Géocodage avec barre de progression améliorée
    st.markdown("---")
    st.subheader("🌍 Géocodage en cours...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    success_counter = st.empty()
    
    coordinates = []
    geocoding_success = []
    total_addresses = len(df_clean)
    success_count = 0
    
    # Estimation du temps
    estimated_time = total_addresses * 1.5  # 1.5 secondes par adresse en moyenne
    start_time = time.time()
    
    for i, address in enumerate(df_clean['adresse_complete']):
        progress = (i + 1) / total_addresses
        progress_bar.progress(progress)
        
        # Calcul du temps restant
        if i > 0:
            elapsed_time = time.time() - start_time
            avg_time_per_address = elapsed_time / i
            remaining_time = avg_time_per_address * (total_addresses - i)
            time_str = f" - Temps restant: {remaining_time:.0f}s"
        else:
            time_str = ""
        
        status_text.text(f"Géocodage : {i+1}/{total_addresses}{time_str}")
        
        lat, lon, success = geocode_address(address)
        coordinates.append((lat, lon))
        geocoding_success.append(success)
        
        if success:
            success_count += 1
        
        success_counter.text(f"✅ Succès: {success_count}/{i+1} ({success_count/(i+1)*100:.1f}%)")
        
        # Délai adaptatif pour éviter le rate limiting
        if i < total_addresses - 1:  # Pas de délai pour la dernière adresse
            time.sleep(1.2)  # Délai légèrement réduit
    
    # Ajout des coordonnées
    df_clean[['lat', 'lon']] = pd.DataFrame(coordinates)
    df_clean['geocoding_success'] = geocoding_success
    
    # Séparation des adresses géocodées/non géocodées
    df_geocoded = df_clean[df_clean['geocoding_success'] == True].dropna(subset=['lat', 'lon']).reset_index(drop=True)
    df_failed = df_clean[df_clean['geocoding_success'] == False].reset_index(drop=True)
    
    st.markdown("---")
    st.subheader("📍 Résultats du géocodage")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("✅ Succès", len(df_geocoded))
    with col2:
        st.metric("❌ Échecs", len(df_failed))
    with col3:
        st.metric("📊 Taux de succès", f"{len(df_geocoded)/len(df_clean)*100:.1f}%")
    
    if len(df_failed) > 0:
        st.warning("⚠️ Adresses non géocodées :")
        st.dataframe(df_failed[[address_col, postal_col, city_col]], use_container_width=True)
    
    if len(df_geocoded) < 2:
        st.error("❌ Pas assez d'adresses géocodées pour créer une tournée")
        st.stop()
    
    # Filtrage intelligent des adresses
    st.markdown("---")
    st.subheader("🎯 Sélection du secteur de livraison")
    
    with st.spinner("Analyse du secteur et sélection des adresses..."):
        df_sector, df_out_sector = filter_addresses_by_proximity(df_geocoded, max_radius)
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("🎯 Adresses dans le secteur", len(df_sector))
    with col2:
        st.metric("🚫 Adresses hors secteur", len(df_out_sector))
    
    if len(df_out_sector) > 0:
        st.warning("🚨 Adresses exclues (trop éloignées du secteur principal) :")
        st.dataframe(df_out_sector[[address_col, postal_col, city_col]], use_container_width=True)
    
    if len(df_sector) < 2:
        st.error("❌ Pas assez d'adresses dans le secteur pour créer une tournée")
        st.stop()
    
    # Optimisation de la tournée
    st.markdown("---")
    st.subheader("🎯 Optimisation de la tournée")
    
    with st.spinner("Calcul de l'itinéraire optimal depuis le centre-ville..."):
        optimal_order = optimize_delivery_route_from_center(df_sector)
        df_optimized = df_sector.loc[optimal_order].reset_index(drop=True)
    
    # Calcul des statistiques
    total_distance, avg_distance, max_distance = calculate_route_stats(df_optimized)
    estimated_time = total_distance * 3 + len(df_optimized) * 5  # 3 min/km + 5 min/arrêt
    
    # Calcul des métriques de qualité
    route_quality = calculate_route_quality_metrics(df_optimized)
    
    # Affichage des résultats
    st.markdown("---")
    st.subheader("📊 Résultats de l'optimisation")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("🏠 Livraisons", len(df_optimized))
    with col2:
        st.metric("📏 Distance totale", f"{total_distance:.1f} km")
    with col3:
        st.metric("⏱️ Temps estimé", f"{estimated_time:.0f} min")
    with col4:
        st.metric("📍 Distance moy/étape", f"{avg_distance:.1f} km")
    
    # Métriques de qualité avancées
    if route_quality:
        st.markdown("### 🎯 Métriques de qualité de la route")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("⚡ Efficacité", f"{route_quality.get('efficiency', 0):.2%}")
        with col2:
            st.metric("📊 Régularité", f"{1 - route_quality.get('coefficient_variation', 1):.2%}")
        with col3:
            st.metric("🎯 Compacité", f"{route_quality.get('compactness', 0):.1f} km")
        with col4:
            quality_score = route_quality.get('quality_score', 0)
            st.metric("⭐ Score qualité", f"{quality_score:.2f}")
    
    # Identification du point de départ
    center_idx = find_center_city_point(df_optimized)
    st.info(f"🎯 **Point de départ** : {df_optimized.iloc[0][address_col]} (Centre-ville détecté)")
    
    # Alertes de qualité
    if route_quality:
        if route_quality.get('max_distance', 0) > 15:
            st.warning(f"⚠️ Distance maximale entre étapes élevée: {route_quality['max_distance']:.1f} km")
        if route_quality.get('efficiency', 1) < 0.3:
            st.warning("⚠️ Efficacité de la route faible - considérez diviser en plusieurs tournées")
        if route_quality.get('quality_score', 0) > 0.7:
            st.success("✅ Excellente qualité de route optimisée!")
    
    # Tableau optimisé
    st.subheader("🗂️ Itinéraire de livraison optimisé")
    
    # Préparation de l'affichage
    df_display = df_optimized.copy()
    df_display.insert(0, 'Ordre', range(1, len(df_display) + 1))
    
    # Calcul des distances entre étapes
    distances_etapes = []
    temps_cumule = 0
    
    for i in range(len(df_display)):
        if i == 0:
            distances_etapes.append("🏁 Départ")
        else:
            coord1 = (df_display.iloc[i-1]['lat'], df_display.iloc[i-1]['lon'])
            coord2 = (df_display.iloc[i]['lat'], df_display.iloc[i]['lon'])
            dist = geodesic(coord1, coord2).kilometers
            temps_cumule += dist * 3 + 5  # 3 min/km + 5 min arrêt
            distances_etapes.append(f"{dist:.1f} km")
    
    df_display['Distance'] = distances_etapes
    
    # Colonnes à afficher
    display_columns = ['Ordre', address_col, postal_col, city_col, 'Distance']
    
    st.dataframe(
        df_display[display_columns],
        use_container_width=True,
        hide_index=True
    )
    
    # Export Excel amélioré
    st.markdown("---")
    st.subheader("💾 Téléchargement")
    
    # Préparation de l'export avec métadonnées
    export_data = df_display.copy()
    
    # Ajout des coordonnées pour GPS
    export_data['Latitude'] = df_optimized['lat']
    export_data['Longitude'] = df_optimized['lon']
    
    # Ajout des informations de temps
    temps_cumule = []
    for i in range(len(df_display)):
        if i == 0:
            temps_cumule.append(0)
        else:
            coord1 = (df_display.iloc[i-1]['lat'], df_display.iloc[i-1]['lon'])
            coord2 = (df_display.iloc[i]['lat'], df_display.iloc[i]['lon'])
            dist = geodesic(coord1, coord2).kilometers
            temps_etape = dist * 3 + 5  # 3 min/km + 5 min arrêt
            temps_cumule.append(temps_cumule[-1] + temps_etape)
    
    export_data['Temps_cumule_min'] = temps_cumule
    
    # Création du fichier Excel avec formatage
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Feuille principale - Tournée optimisée
        export_data.to_excel(writer, index=False, sheet_name='Tournée_optimisée')
        
        # Feuille des statistiques
        stats_data = {
            'Métrique': [
                'Nombre total d\'adresses',
                'Adresses géocodées',
                'Adresses dans le secteur',
                'Adresses hors secteur',
                'Adresses non géocodées',
                'Distance totale (km)',
                'Temps estimé (min)',
                'Distance moyenne par étape (km)',
                'Efficacité du géocodage (%)',
                'Rayon du secteur utilisé (km)'
            ],
            'Valeur': [
                len(df_clean),
                len(df_geocoded),
                len(df_sector),
                len(df_out_sector),
                len(df_failed),
                f"{total_distance:.1f}",
                f"{estimated_time:.0f}",
                f"{avg_distance:.1f}",
                f"{len(df_geocoded)/len(df_clean)*100:.1f}",
                f"{max_radius:.1f}"
            ]
        }
        pd.DataFrame(stats_data).to_excel(writer, index=False, sheet_name='Statistiques')
        
        # Feuille des échecs de géocodage
        if len(df_failed) > 0:
            df_failed_export = df_failed.copy()
            df_failed_export['Raison'] = 'Géocodage échoué'
            df_failed_export.to_excel(writer, index=False, sheet_name='Échecs_géocodage')
        
        # Feuille des adresses hors secteur
        if len(df_out_sector) > 0:
            df_out_sector_export = df_out_sector.copy()
            # Calculer la distance du centre pour chaque adresse hors secteur
            center_idx = find_center_city_point(df_sector)
            if center_idx in df_sector.index:
                center_coords = (df_sector.at[center_idx, 'lat'], df_sector.at[center_idx, 'lon'])
                distances_center = []
                for _, row in df_out_sector.iterrows():
                    point_coords = (row['lat'], row['lon'])
                    dist = geodesic(center_coords, point_coords).kilometers
                    distances_center.append(f"{dist:.1f}")
                df_out_sector_export['Distance_du_centre_km'] = distances_center
            
            df_out_sector_export.to_excel(writer, index=False, sheet_name='Hors_secteur')
        
        # Formatage des feuilles
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Formatage de la feuille principale
            ws_main = writer.sheets['Tournée_optimisée']
            
            # En-têtes en gras
            for col in range(1, ws_main.max_column + 1):
                ws_main.cell(row=1, column=col).font = Font(bold=True)
                ws_main.cell(row=1, column=col).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Ajustement automatique des largeurs de colonnes
            for col in range(1, ws_main.max_column + 1):
                column_letter = get_column_letter(col)
                ws_main.column_dimensions[column_letter].width = 15
            
            # Formatage de la feuille des statistiques
            if 'Statistiques' in writer.sheets:
                ws_stats = writer.sheets['Statistiques']
                for col in range(1, ws_stats.max_column + 1):
                    ws_stats.cell(row=1, column=col).font = Font(bold=True)
                    ws_stats.cell(row=1, column=col).fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                    column_letter = get_column_letter(col)
                    ws_stats.column_dimensions[column_letter].width = 25
                    
        except ImportError:
            # Si openpyxl.styles n'est pas disponible, continuer sans formatage
            pass
    
    output.seek(0)
    
    # Nom du fichier avec informations
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tournee_optimisee_{len(df_optimized)}adresses_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Télécharger la tournée optimisée",
        data=output,
        file_name=filename,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type="primary",
        use_container_width=True
    )
    
    # Résumé final
    st.success("✅ Tournée optimisée avec succès !")
    st.markdown(f"**🎯 {len(df_optimized)} livraisons** dans le secteur sur **{total_distance:.1f} km** en **{estimated_time:.0f} minutes**")
    
    # Conseils
    st.markdown("---")
    st.subheader("💡 Conseils d'optimisation")
    st.info("""
    ✅ **Itinéraire optimisé** : Départ du centre-ville vers la périphérie
    ✅ **Secteur cohérent** : Adresses trop éloignées automatiquement exclues
    ✅ **Ordre logique** : Parcours géographiquement cohérent
    ✅ **Temps réaliste** : Inclut temps de conduite + temps d'arrêt
    """)

# Footer
st.markdown("---")
st.markdown("🚛 **Organisateur de Tournées Automatique** - Optimisation intelligente pour vos livraisons")
st.markdown("*Géocodage automatique • Filtrage intelligent • Optimisation depuis centre-ville • Export Excel*")
